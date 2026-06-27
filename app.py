# app.py  (Cloud Run 배포 버전)
import asyncio, uuid, subprocess
import csv
import shutil
from datetime import datetime
from pathlib import Path
from fastapi import FastAPI, UploadFile, Form, Request, BackgroundTasks
from fastapi import Body
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import pandas as pd
import io
import warnings

# Silence harmless openpyxl default style warning for some workbooks
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module=r"openpyxl\.styles\.stylesheet"
)

app = FastAPI()
BASE = Path("data"); BASE.mkdir(exist_ok=True, parents=True)
REPORTS_DIR = Path("data/reports"); REPORTS_DIR.mkdir(exist_ok=True, parents=True)
# Serve static assets (mascot video)
app.mount("/asset", StaticFiles(directory="asset"), name="asset")

from fastapi.responses import HTMLResponse

# Slot label order for pivot sorting (output uses uppercase).
SLOT_LABELS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def normalize_student_id(val):
    """Normalize student id to a consistent string (strip decimals, whitespace)."""
    try:
        return str(int(float(val)))
    except Exception:
        try:
            return str(val).strip()
        except Exception:
            return ""

def build_unassigned_lookup(df_asg: pd.DataFrame):
    """Return mapping: subject -> list of {{student_id, name, choices}} for unassigned students."""
    try:
        if df_asg is None or df_asg.empty:
            return {}
        df = df_asg.copy()
        df["student_id_norm"] = df.get("student_id", "").apply(normalize_student_id)
        # Build full choice list per student
        choices_map = {}
        for sid, grp in df.groupby("student_id_norm"):
            choices = []
            for _, row in grp.iterrows():
                subj = str(row.get("subject", "")).strip()
                status = str(row.get("status", "")).strip()
                slot_val = row.get("slot", "")
                slot_s = "" if pd.isna(slot_val) else str(slot_val).strip()
                choices.append({"subject": subj, "status": status, "slot": slot_s})
            choices_map[sid] = choices
        # Collect unassigned per subject
        ua_map = {}
        df_ua = df[df.get("status") == "unassigned"]
        for subj, grp in df_ua.groupby("subject"):
            students = []
            for _, row in grp.iterrows():
                sid = row.get("student_id_norm", "")
                name_val = row.get("name", "")
                name = "" if pd.isna(name_val) else str(name_val).strip()
                students.append({
                    "student_id": sid,
                    "name": name,
                    "choices": choices_map.get(sid, []),
                })
            ua_map[str(subj)] = students
        return ua_map
    except Exception:
        return {}

def build_pivot(assignments_csv: Path):
    try:
        if not assignments_csv.exists():
            return None
        df_all = pd.read_csv(assignments_csv)
        if df_all.empty:
            return None
        # Normalize slot labels for consistent display/sorting (support old lowercase outputs too).
        if "slot" in df_all.columns:
            df_all["slot"] = (
                df_all["slot"]
                .fillna("")
                .astype(str)
                .map(lambda s: s.strip().upper())
            )
        # Subject universe - preserve order from assignments.csv (first occurrence)
        subjects = df_all["subject"].astype(str).unique().tolist()
        # Assigned data for section counts
        df_asg = df_all[df_all.get("status") == "assigned"].copy()
        if not df_asg.empty:
            parts = df_asg["section_label"].astype(str).str.rsplit("_", n=2, expand=True)
            secnum = pd.to_numeric(parts[2], errors="coerce").fillna(0).astype(int)
            df_asg["slot_section"] = df_asg["slot"].astype(str) + "-" + secnum.astype(str)
            ct_sec = pd.crosstab(df_asg["subject"], df_asg["slot_section"]).astype(int)
            sec_cols = ct_sec.columns.tolist()
        else:
            ct_sec = pd.DataFrame()
            sec_cols = []
        # Sort section columns by slot then number
        def sort_key(col: str):
            try:
                slot, num = col.split("-", 1)
                slot_norm = str(slot).strip().upper()
                si = SLOT_LABELS.index(slot_norm) if slot_norm in SLOT_LABELS else 999
                ni = int(num)
                return (si, ni)
            except Exception:
                return (999, 999)
        sec_cols_sorted = sorted(sec_cols, key=sort_key)
        # Demand and unassigned counts
        demand = df_all.groupby("subject").size()
        unassigned = df_all[df_all.get("status") == "unassigned"].groupby("subject").size()
        # Build table rows
        table = pd.DataFrame(index=subjects)
        for c in sec_cols_sorted:
            table[c] = ct_sec.get(c, pd.Series(0, index=ct_sec.index)).reindex(subjects, fill_value=0)
        # Total = demand (assigned + unassigned)
        table["Total"] = demand.reindex(subjects, fill_value=0)
        # Order columns: Total then sections, preserve original subject order
        ordered = ["Total"] + sec_cols_sorted
        # Build JSON-friendly structure
        columns = ["Subject"] + ordered
        rows = []
        row_meta = {}
        for subj in table.index.tolist():
            vals = [int(table.at[subj, c]) for c in ordered]
            rows.append([subj] + vals)
            ua = int(unassigned.get(subj, 0))
            assigned = int(vals[0] - ua) if vals else 0
            row_meta[str(subj)] = {"unassigned": max(ua, 0), "assigned": max(assigned, 0)}
        # Build group headers from section columns
        groups = []
        last_slot = None
        secs: list[int] = []
        for col in sec_cols_sorted:
            try:
                slot, num = col.split("-", 1)
                n = int(num)
            except Exception:
                slot, n = col, 0
            if last_slot is None or slot != last_slot:
                if last_slot is not None:
                    groups.append({"slot": last_slot, "sections": secs})
                last_slot = slot
                secs = [n]
            else:
                secs.append(n)
        if last_slot is not None:
            groups.append({"slot": last_slot, "sections": secs})
        # Slot-level metadata: total opened sections per slot (from sections_plan if available)
        slot_meta = {}
        try:
            sp_path = assignments_csv.parent / "sections_plan.csv"
            if sp_path.exists():
                sp_df = pd.read_csv(sp_path)
                if "slot" in sp_df.columns:
                    sp_df["slot"] = (
                        sp_df["slot"]
                        .fillna("")
                        .astype(str)
                        .map(lambda s: s.strip().upper())
                    )
                agg = sp_df.groupby("slot")["num_sections"].sum().to_dict()
                for sl, cnt in agg.items():
                    slot_meta[str(sl)] = {"sections_open": int(cnt)}
            else:
                # Fallback: approximate by counting distinct (subject, slot, section_no) seen in assignments
                seen = {}
                for col in sec_cols_sorted:
                    try:
                        slot, num = col.split("-", 1)
                        seen.setdefault(slot, set()).add(int(num))
                    except Exception:
                        continue
                for sl, nums in seen.items():
                    slot_meta[str(sl)] = {"sections_open": int(len(nums))}
        except Exception:
            pass
        return {"columns": columns, "rows": rows, "groups": groups, "row_meta": row_meta, "slot_meta": slot_meta}
    except Exception:
        return None


def build_export_xlsx(assignments_csv: Path, output_xlsx: Path, subject_order: list[str] | None = None) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font

    if not assignments_csv.exists():
        raise FileNotFoundError(str(assignments_csv))

    df = pd.read_csv(assignments_csv)
    if df.empty:
        raise ValueError("assignments.csv is empty")

    # Normalize for stable keys
    df["student_id_norm"] = df.get("student_id", "").apply(normalize_student_id)
    df["name_norm"] = df.get("name", "").fillna("").astype(str).map(lambda s: s.strip())
    df["subject_norm"] = df.get("subject", "").fillna("").astype(str).map(lambda s: s.strip())
    df["status_norm"] = df.get("status", "").fillna("").astype(str).map(lambda s: s.strip())
    df["slot_norm"] = df.get("slot", "").fillna("").astype(str).map(lambda s: s.strip().upper())
    df["section_label_norm"] = df.get("section_label", "").fillna("").astype(str).map(lambda s: s.strip())

    # Student roster (id -> name)
    roster: dict[str, str] = {}
    for _, row in df.iterrows():
        sid = row["student_id_norm"]
        if not sid:
            continue
        name = row["name_norm"]
        if sid not in roster:
            roster[sid] = name
        elif not roster[sid] and name:
            roster[sid] = name

    # Subject universe
    subjects_in_data = df["subject_norm"].tolist()
    # Preserve first-seen order
    seen = set()
    subjects_first = []
    for s in subjects_in_data:
        if not s or s in seen:
            continue
        seen.add(s)
        subjects_first.append(s)

    ordered: list[str] = []
    if subject_order:
        subjects_set = set(subjects_first)
        for s in subject_order:
            ss = str(s).strip()
            if ss and ss in subjects_set and ss not in ordered:
                ordered.append(ss)
    for s in subjects_first:
        if s not in ordered:
            ordered.append(s)

    # Assignment mapping: (sid, subject) -> (slot, division_no) or None when unassigned.
    asg_map: dict[tuple[str, str], tuple[str, int | None] | None] = {}

    def _parse_division(label: str) -> int | None:
        if not label:
            return None
        try:
            tail = label.rsplit("_", 1)[-1]
            return int(tail)
        except Exception:
            return None

    for _, row in df.iterrows():
        sid = row["student_id_norm"]
        subj = row["subject_norm"]
        if not sid or not subj:
            continue
        status = row["status_norm"]
        if status != "assigned":
            asg_map.setdefault((sid, subj), None)
            continue
        slot = row["slot_norm"]
        div = _parse_division(row["section_label_norm"])
        if slot:
            asg_map[(sid, subj)] = (slot, div)
        else:
            asg_map[(sid, subj)] = None

    # Sort students by numeric id when possible
    def sid_sort_key(sid: str):
        try:
            return (0, int(sid))
        except Exception:
            return (1, sid)

    student_ids = sorted(roster.keys(), key=sid_sort_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "배정결과"

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Two-row header: subject merged across 2 columns
    ws.cell(row=1, column=1, value="학번").font = header_font
    ws.cell(row=1, column=2, value="이름").font = header_font
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=2).alignment = center

    col = 3
    for subj in ordered:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        c = ws.cell(row=1, column=col, value=subj)
        c.font = header_font
        c.alignment = center
        c1 = ws.cell(row=2, column=col, value="섹션")
        c2 = ws.cell(row=2, column=col + 1, value="분반")
        c1.font = header_font
        c2.font = header_font
        c1.alignment = center
        c2.alignment = center
        col += 2

    # Freeze headers and id/name columns
    ws.freeze_panes = "C3"

    # Set widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    for i in range(3, 3 + 2 * len(ordered)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10

    # Body rows
    r = 3
    for sid in student_ids:
        ws.cell(row=r, column=1, value=sid)
        ws.cell(row=r, column=2, value=roster.get(sid, ""))
        c = 3
        for subj in ordered:
            val = asg_map.get((sid, subj))
            if val is None:
                # unassigned or not chosen -> blank cells
                c += 2
                continue
            slot, div = val
            ws.cell(row=r, column=c, value=slot if slot else "")
            ws.cell(row=r, column=c + 1, value=div if div and div > 0 else "")
            c += 2
        r += 1

    # Align all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment = Alignment(vertical="center")

    wb.save(output_xlsx)

@app.get("/", response_class=HTMLResponse)
def index():
    return f"""
    <!doctype html>
    <html lang=\"ko\">
    <head>
      <meta charset=\"utf-8\" />
      <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
      <title>고교학점제 이동반 편성 프로그램</title>
      <script src=\"https://cdn.tailwindcss.com\"></script>
      <style>
        :root {{
          --sage-50: #F6F8F6;
          --sage-100: #E9F0EC;
          --sage-200: #D5E2DA;
          --sage-300: #C0D2C6;
          --sage-400: #A8C0AE;
          --sage-500: #8FAA96;
          --sage-600: #7A957F;
          --sage-700: #647A68;
        }}
        .bg-sage-50 {{ background-color: var(--sage-50); }}
        .bg-sage-100 {{ background-color: var(--sage-100); }}
        .bg-sage-500 {{ background-color: var(--sage-500); }}
        .bg-sage-600 {{ background-color: var(--sage-600); }}
        .hover\\:bg-sage-200:hover {{ background-color: var(--sage-200); }}
        .hover\\:bg-sage-700:hover {{ background-color: var(--sage-700); }}
        .text-sage-700 {{ color: var(--sage-700); }}
        .text-sage-800 {{ color: var(--sage-700); }}
        .ring-sage-200 {{ --tw-ring-color: var(--sage-200); }}
        .ring-sage-300 {{ --tw-ring-color: var(--sage-300); }}
        .focus\\:ring-sage-300:focus {{ --tw-ring-color: var(--sage-300); }}
        .file\\:bg-sage-100 {{ background-color: var(--sage-100); }}
        .file\\:text-sage-800 {{ color: var(--sage-700); }}

        /* Pivot drag-and-drop insertion indicator */
        .drop-line-before td {{ box-shadow: inset 0 3px 0 var(--sage-600); }}
        .drop-line-after td {{ box-shadow: inset 0 -3px 0 var(--sage-600); }}
      </style>
    </head>
    <body class=\"bg-sage-50 text-stone-800\">
      <div class=\"mx-auto max-w-5xl p-6\">
        <header class=\"mb-6\">
          <h1 class=\"text-2xl font-semibold tracking-tight text-stone-900\">고교학점제 이동반 편성 프로그램</h1>
          <p class=\"text-sm text-stone-500\">OR-Tools 기반 미배정 최소화 모델</p>
        </header>

        <section class=\"bg-white border border-stone-200 rounded-xl shadow-sm\">
          <div class=\"p-5\">
            <form id=\"run-form\" enctype=\"multipart/form-data\">
              <div class=\"mb-4\">
                <label class=\"block text-sm font-medium text-stone-700 mb-1\">고교학점제 신규 수강신청 프로그램의 수강신청-신청결과-템플릿다운로드 양식(.xlsx) 업로드</label>
                <input class=\"block w-full text-sm file:mr-4 file:py-2 file:px-3 file:rounded-md file:border-0 file:text-sm file:font-medium file:bg-sage-100 file:text-sage-800 hover:file:bg-sage-200 border border-stone-300 rounded-lg p-2 bg-white\" type=\"file\" name=\"xlsx\" accept=\".xlsx\" required />
              </div>

              
              <div id=\"group-row\" class=\"mb-4 hidden\"> 
                <label class=\"block text-sm font-medium text-stone-700 mb-1\">선택그룹 (여러 개 선택 가능)</label>
                <select id=\"group-select\" name=\"group\" multiple size=\"4\" class=\"w-full rounded-lg border border-stone-300 p-2 bg-white\"></select>
                <p class=\"mt-1 text-xs text-stone-500\">여러 그룹을 포함하려면 Ctrl/Command 또는 Shift 키를 사용해 선택하세요. 아무 것도 선택하지 않으면 모든 그룹이 포함됩니다.</p>
              </div>

              <div id=\"inspect-wrap\" class=\"mb-4 hidden\">
                <div class=\"rounded-lg border border-sage-200 bg-sage-50 p-3 text-sm\" id=\"inspect-body\"></div>
              </div>

              <div id=\"subject-preview-wrap\" class=\"mb-4 hidden rounded-lg border border-stone-200 bg-white shadow-inner\">
                <div class=\"flex flex-col gap-2 border-b border-stone-200 p-3 md:flex-row md:items-center md:justify-between\">
                  <div>
                    <div class=\"text-sm font-medium text-stone-900\">과목별 추천 반수</div>
                    <p class=\"text-xs text-stone-500\">현재 정원 기준 추천 반수가 자동으로 고정됩니다. 값을 수정하거나 비워두면 해당 과목만 조정됩니다.</p>
                  </div>
                  <div class=\"flex gap-2\">
                    <button id=\"apply-section-recommend\" type=\"button\" class=\"rounded-md border border-sage-300 px-3 py-1 text-sm text-sage-700 hover:bg-sage-50\">추천값 다시 적용</button>
                    <button id=\"release-all-sections\" type=\"button\" class=\"rounded-md border border-stone-300 px-3 py-1 text-sm text-stone-600 hover:bg-stone-50\">모두 자동</button>
                  </div>
                </div>
                <div class=\"overflow-x-auto\">
                  <table class=\"min-w-full divide-y divide-stone-200 text-sm\">
                    <thead class=\"bg-sage-100\">
                      <tr>
                        <th class=\"px-3 py-2 text-left font-medium text-stone-700\">과목</th>
                        <th class=\"px-3 py-2 text-right font-medium text-stone-700\">선택인원</th>
                        <th class=\"px-3 py-2 text-right font-medium text-stone-700\">권장 반수</th>
                        <th class=\"px-3 py-2 text-left font-medium text-stone-700\">개설 반수</th>
                        <th class=\"px-3 py-2 text-right font-medium text-stone-700\">평균 학생수</th>
                      </tr>
                    </thead>
                    <tbody id=\"subject-preview-body\" class=\"divide-y divide-stone-100 bg-white\"></tbody>
                  </table>
                </div>
                <div class=\"border-t border-stone-200 p-3 text-xs text-stone-500\">
                  값이 비어 있으면 해당 과목은 자동으로 결정됩니다. 입력값을 바꾸면 그 숫자로 분반 수가 고정됩니다.
                </div>
              </div>

              <div class=\"grid grid-cols-1 md:grid-cols-3 gap-3\">
                <div>
                  <label class=\"block text-sm text-stone-700 mb-1\">선택수</label>
                  <input class=\"w-full rounded-lg border border-stone-300 p-2 focus:outline-none focus:ring-2 focus:ring-sage-300\" type=\"number\" name=\"slots\" min=\"1\" value=\"4\" />
                </div>
                <div>
                  <label class=\"block text-sm text-stone-700 mb-1\">섹션당 교실수</label>
                  <input class=\"w-full rounded-lg border border-stone-300 p-2 focus:outline-none focus:ring-2 focus:ring-sage-300\" type=\"number\" name=\"rooms\" min=\"1\" value=\"7\" />
                </div>
                <div>
                  <label class=\"block text-sm text-stone-700 mb-1\">섹션당 추가가능교실수</label>
                  <input class=\"w-full rounded-lg border border-stone-300 p-2 focus:outline-none focus:ring-2 focus:ring-sage-300\" type=\"number\" name=\"extra\" min=\"0\" value=\"1\" />
                </div>
                <div>
                  <label class=\"block text-sm text-stone-700 mb-1\">교실정원</label>
                  <input class=\"w-full rounded-lg border border-stone-300 p-2 focus:outline-none focus:ring-2 focus:ring-sage-300\" type=\"number\" name=\"cap\" min=\"1\" value=\"28\" />
                </div>
                <div>
                  <label class=\"block text-sm text-stone-700 mb-1\">교실최대정원</label>
                  <input class=\"w-full rounded-lg border border-stone-300 p-2 focus:outline-none focus:ring-2 focus:ring-sage-300\" type=\"number\" name=\"maxcap\" min=\"1\" value=\"30\" />
                </div>
                <div>
                  <label class=\"block text-sm text-stone-700 mb-1\">전체 섹션에서 추가가능한 교실수</label>
                  <input class=\"w-full rounded-lg border border-stone-300 p-2 focus:outline-none focus:ring-2 focus:ring-sage-300\" type=\"number\" name=\"extra_total\" min=\"0\" placeholder=\"무제한이면 비워두기\" />
                </div>
              </div>
              <div class=\"mt-5 flex items-center gap-3\">
                <button id=\"run-submit\" type=\"submit\" class=\"px-4 py-2 rounded-lg bg-sage-600 text-white hover:bg-sage-700 focus:ring-2 focus:ring-sage-300\">실행</button>
                <a class=\"text-sm text-stone-500 hover:text-stone-700\" href=\"/docs\">API 문서 보기</a>
              </div>
            </form>
          </div>
        </section>

        <section class=\"mt-6 bg-white border border-stone-200 rounded-xl shadow-sm\">
          <div class=\"p-5\">
            <div class=\"flex items-center justify-between mb-3\">
              <h2 class=\"text-lg font-medium text-stone-900\">상태 & 결과</h2>
              <span id=\"status-badge\" class=\"hidden inline-flex items-center px-2 py-1 text-xs rounded-full bg-sage-50 text-sage-700 ring-1 ring-sage-200\">PENDING</span>
            </div>
            <div id=\"status-note\" class=\"text-sm text-stone-500\">실행하면 여기에서 진행상황과 다운로드 링크가 표시됩니다.</div>
            <div id=\"job-info\" class=\"mt-2 text-sm text-stone-600 hidden\"></div>
            <div id=\"mascot\" class=\"mt-4 hidden flex justify-center items-center gap-4\">
              <video id=\"mascot-video\" src=\"/asset/beori2.mp4\" autoplay loop muted playsinline class=\"w-36 h-36 md:w-40 md:h-40 object-contain rounded-lg shadow-sm ring-1 ring-sage-200\"></video>
              <img id=\"mascot-done\" src=\"/asset/beori2_done.png\" onerror=\"this.onerror=null; this.src='/asset/beori_done2.png';\" alt=\"done\" class=\"hidden w-36 h-36 md:w-40 md:h-40 object-contain rounded-lg shadow-sm ring-1 ring-sage-200\" />
            </div>
              <div id=\"downloads\" class=\"mt-4 space-x-2 hidden\"></div>
            <div id=\"pivot-wrap\" class=\"mt-6 hidden overflow-x-auto\">
              <table id=\"pivot-table\" class=\"min-w-full text-sm\"></table>
            </div>

            <!-- Unassigned Detail Modal -->
            <div id=\"unassigned-modal\" class=\"fixed inset-0 bg-black bg-opacity-50 hidden z-50 flex items-center justify-center\">
              <div class=\"bg-white rounded-lg p-6 w-[720px] max-w-full mx-4 shadow-lg\">
                <div class=\"flex justify-between items-center mb-4\">
                  <div>
                    <h3 id=\"unassigned-title\" class=\"text-lg font-medium text-stone-900\">미배정 학생</h3>
                    <p id=\"unassigned-count\" class=\"text-xs text-stone-500\"></p>
                  </div>
                  <button id=\"close-unassigned\" class=\"text-stone-400 hover:text-stone-600\">
                    <svg class=\"w-6 h-6\" fill=\"none\" stroke=\"currentColor\" viewBox=\"0 0 24 24\">
                      <path stroke-linecap=\"round\" stroke-linejoin=\"round\" stroke-width=\"2\" d=\"M6 18L18 6M6 6l12 12\"></path>
                    </svg>
                  </button>
                </div>
                <div id=\"unassigned-content\" class=\"max-h-64 overflow-y-auto border border-stone-200 rounded-lg bg-sage-50 p-3 text-sm text-stone-700\">
                  <div class=\"text-stone-500\">데이터를 불러오는 중...</div>
                </div>
              </div>
            </div>

            <!-- Constraint Settings Modal -->
            <div id=\"constraint-modal\" class=\"fixed inset-0 bg-black bg-opacity-50 hidden z-50 flex items-center justify-center\">
              <div class=\"bg-white rounded-lg p-6 w-96 max-w-full mx-4\">
                <div class=\"flex justify-between items-center mb-4\">
                  <h3 id=\"modal-subject-title\" class=\"text-lg font-medium text-stone-900\">과목 제약 조건 설정</h3>
                  <button id=\"close-modal\" class=\"text-stone-400 hover:text-stone-600\">
                    <svg class=\"w-6 h-6\" fill=\"none\" stroke=\"currentColor\" viewBox=\"0 0 24 24\">
                      <path stroke-linecap=\"round\" stroke-linejoin=\"round\" stroke-width=\"2\" d=\"M6 18L18 6M6 6l12 12\"></path>
                    </svg>
                  </button>
                </div>
                <div class=\"space-y-4\">
                  <div>
                  <label class=\"block text-sm font-medium text-stone-700 mb-1\">동시간에 수업 가능한 교사 수</label>
                  <input id=\"max-per-slot\" type=\"number\" min=\"0\" placeholder=\"제한 없음\" class=\"w-full rounded-lg border border-stone-300 p-2 focus:outline-none focus:ring-2 focus:ring-sage-300\" />
                  <p class=\"mt-1 text-xs text-stone-500\">각 시간대에 이 과목이 개설할 수 있는 최대 반 수</p>
                </div>
                  <div class=\"flex justify-end space-x-3 pt-4\">
                    <button id=\"clear-constraints\" class=\"px-4 py-2 text-stone-600 hover:text-stone-800\">제약 해제</button>
                    <button id=\"save-constraints\" class=\"px-4 py-2 bg-sage-600 text-white rounded-lg hover:bg-sage-700\">저장</button>
                  </div>
                </div>
              </div>
            </div>
            <pre id=\"error-box\" class=\"mt-4 hidden text-sm text-rose-700 bg-rose-50 border border-rose-200 rounded-lg p-3 whitespace-pre-wrap\"></pre>
            
          </div>
        </section>

        <footer class=\"mt-8 border-t border-stone-200 pt-6\">
          <div class=\"text-xs text-stone-500 space-y-3\">
            <div class=\"flex flex-col md:flex-row md:items-center md:justify-between gap-3\">
              <div class=\"space-y-1\">
                <p class=\"font-medium text-stone-700\">고교학점제 이동반 편성 프로그램</p>
                <p class=\"text-stone-500\">개발: <a href=\"https://namgungyeon.tistory.com/138\" target=\"_blank\" rel=\"noopener noreferrer\" class=\"text-sage-700 hover:text-sage-800 underline\">남궁연</a> (설악고등학교 교사)</p>
              </div>

              <div class=\"flex items-center gap-2\">
                <a href=\"https://creativecommons.org/licenses/by-nc/4.0/\" target=\"_blank\" rel=\"noopener noreferrer\" title=\"CC BY-NC 4.0 License\">
                  <img src=\"https://img.shields.io/badge/License-CC%20BY--NC%204.0-lightgrey.svg\" alt=\"CC BY-NC 4.0 License\" class=\"h-5\" />
                </a>
              </div>
            </div>

            <p class=\"text-stone-400 text-[10px]\">이 프로그램은 교육 및 연구 목적으로 자유롭게 사용 가능합니다. 상업적 이용은 불가합니다.</p>
          </div>
        </footer>
      </div>

      <script>
        const form = document.getElementById('run-form');
        const submitBtn = document.getElementById('run-submit');
        const statusBadge = document.getElementById('status-badge');
        const statusNote = document.getElementById('status-note');
        const downloads = document.getElementById('downloads');
        const jobInfo = document.getElementById('job-info');
        const errBox = document.getElementById('error-box');
        const pivotWrap = document.getElementById('pivot-wrap');
        const pivotTable = document.getElementById('pivot-table');
        const inputFile = form.querySelector('input[name="xlsx"]');
        const groupRow = document.getElementById('group-row');
        const groupSel = document.getElementById('group-select');
        const mascot = document.getElementById('mascot');
        const mascotVideo = document.getElementById('mascot-video');
        const mascotDone = document.getElementById('mascot-done');
        const inspectWrap = document.getElementById('inspect-wrap');
        const inspectBody = document.getElementById('inspect-body');
        const subjectPreviewWrap = document.getElementById('subject-preview-wrap');
        const subjectPreviewBody = document.getElementById('subject-preview-body');
        const applySectionRecommendBtn = document.getElementById('apply-section-recommend');
        const releaseAllSectionsBtn = document.getElementById('release-all-sections');
        const capInput = form.querySelector('input[name="cap"]');
        const unassignedModal = document.getElementById('unassigned-modal');
        const unassignedTitle = document.getElementById('unassigned-title');
        const unassignedCount = document.getElementById('unassigned-count');
        const unassignedContent = document.getElementById('unassigned-content');
        const closeUnassignedBtn = document.getElementById('close-unassigned');
        let pollTimer = null;
        let pivotData = null;
        let lastInspect = null;
        let sortAsc = false; // total sort direction (desc by default)
        let pivotSortMode = 'subject'; // 'subject' | 'total'
        let subjectSortAsc = true; // subject sort direction (asc by default)
        let runStartMs = null;    // client-side ticking timer
        let tickTimer = null;     // interval handle for elapsed seconds
        let lastProgress = null;  // latest progress payload
        let subjectConstraints = {{}};  // subject name -> {{maxPerSlot}}
        let currentModalSubject = null;  // currently editing subject in modal
        let fixedSectionTargets = {{}};
        const normalizeGroupValue = (val) => (val === undefined || val === null) ? '' : String(val).trim();
        let currentGroupFilter = new Set();
        let subjectCounts = {{}};
        let currentJobId = null;          // latest job id (may be running)
        let latestPivotJobId = null;      // job id tied to the pivot currently rendered
        let rowOrderMode = 'total';       // 'total' | 'manual'
        let manualSubjectOrder = [];      // subject name array
        let pivotDnDInitialized = false;
        let pivotDragSubject = null;
        let pivotDragSubjectKey = null;
        let pivotDropPosition = 'before'; // 'before' | 'after'
        const SUBJECT_ORDER_STORAGE_KEY = 'hscredit_subject_order_v1';

        function loadStoredSubjectOrder() {{
          try {{
            const raw = localStorage.getItem(SUBJECT_ORDER_STORAGE_KEY);
            if (!raw) return [];
            const arr = JSON.parse(raw);
            return Array.isArray(arr) ? arr.map((s) => String(s)) : [];
          }} catch (e) {{
            return [];
          }}
        }}

        function saveStoredSubjectOrder(order) {{
          try {{
            const uniq = [];
            const seen = new Set();
            (Array.isArray(order) ? order : []).forEach((s) => {{
              const v = String(s);
              if (!v || seen.has(v)) return;
              seen.add(v);
              uniq.push(v);
            }});
            localStorage.setItem(SUBJECT_ORDER_STORAGE_KEY, JSON.stringify(uniq));
          }} catch (e) {{
            // ignore
          }}
        }}

        function initPivotDnD() {{
          if (pivotDnDInitialized || !pivotTable) return;
          pivotDnDInitialized = true;

          const clearDragState = () => {{
            pivotDragSubject = null;
            pivotDragSubjectKey = null;
            pivotTable.querySelectorAll('tr.dragging-row').forEach((el) => {{
              el.classList.remove('dragging-row', 'opacity-70', 'ring-2', 'ring-sage-500', 'bg-sage-50');
            }});
            pivotTable.querySelectorAll('.subject-drag-handle.font-semibold').forEach((el) => {{
              el.classList.remove('font-semibold', 'text-sage-700');
            }});
          }};

          const clearDropHighlights = () => {{
            pivotTable.querySelectorAll('tr.drop-line-before, tr.drop-line-after').forEach((el) => {{
              el.classList.remove('drop-line-before', 'drop-line-after');
            }});
          }};

          pivotTable.addEventListener('dragstart', (e) => {{
            const handle = e.target.closest('.subject-drag-handle');
            if (!handle) return;
            const key = handle.getAttribute('data-subject-key');
            if (!key) return;
            pivotDragSubjectKey = key;
            try {{
              pivotDragSubject = decodeURIComponent(key);
            }} catch (err) {{
              pivotDragSubject = key;
            }}
            rowOrderMode = 'manual';
            // Visually mark the row being dragged.
            const row = pivotTable.querySelector(`tr[data-subject-key="${{escapeSelector(key)}}"]`);
            if (row) {{
              row.classList.add('dragging-row', 'opacity-70', 'ring-2', 'ring-sage-500', 'bg-sage-50');
            }}
            handle.classList.add('font-semibold', 'text-sage-700');
            if (e.dataTransfer) {{
              e.dataTransfer.effectAllowed = 'move';
              e.dataTransfer.setData('text/plain', key);
            }}
          }});

          pivotTable.addEventListener('dragover', (e) => {{
            if (!pivotDragSubject) return;
            const row = e.target.closest('tr[data-subject-key]');
            if (!row) return;
            if (row.classList.contains('dragging-row')) return;
            e.preventDefault();
            clearDropHighlights();
            const rect = row.getBoundingClientRect();
            const y = e.clientY - rect.top;
            const before = y < rect.height / 2;
            pivotDropPosition = before ? 'before' : 'after';
            row.classList.add(before ? 'drop-line-before' : 'drop-line-after');
          }});

          pivotTable.addEventListener('drop', (e) => {{
            if (!pivotDragSubject) return;
            const row = e.target.closest('tr[data-subject-key]');
            if (!row) return;
            e.preventDefault();
            clearDropHighlights();
            const dropKey = row.getAttribute('data-subject-key');
            if (!dropKey) return;
            let dropSubject = dropKey;
            try {{
              dropSubject = decodeURIComponent(dropKey);
            }} catch (err) {{
              // keep as-is
            }}
            if (!pivotData || !Array.isArray(pivotData.rows)) return;
            const subjectsNow = pivotData.rows.map((r) => String(r[0]));
            if (!manualSubjectOrder || !manualSubjectOrder.length) {{
              manualSubjectOrder = subjectsNow.slice();
            }}
            // Ensure manual order contains all current subjects (append missing)
            const inOrder = new Set(manualSubjectOrder);
            subjectsNow.forEach((s) => {{
              if (!inOrder.has(s)) manualSubjectOrder.push(s);
            }});

            const from = pivotDragSubject;
            const to = dropSubject;
            if (!from || !to || from === to) return;
            manualSubjectOrder = manualSubjectOrder.filter((s) => s !== from);
            const toIndex = manualSubjectOrder.indexOf(to);
            const baseIndex = Math.max(0, toIndex);
            const insertAt = pivotDropPosition === 'after' ? baseIndex + 1 : baseIndex;
            manualSubjectOrder.splice(insertAt, 0, from);
            saveStoredSubjectOrder(manualSubjectOrder);
            drawPivot();
          }});

          pivotTable.addEventListener('dragend', () => {{
            clearDropHighlights();
            clearDragState();
          }});
        }}

        function applyGroupFilter(values) {{
          const normalized = new Set((Array.isArray(values) ? values : []).map((v) => normalizeGroupValue(v)).filter(Boolean));
          if (groupSel) {{
            Array.from(groupSel.options || []).forEach((opt) => {{
              const val = normalizeGroupValue(opt.value);
              opt.selected = normalized.has(val);
            }});
          }}
          currentGroupFilter = normalized;
        }}

        if (groupSel) {{
          groupSel.addEventListener('change', () => {{
            const selected = Array.from(groupSel.selectedOptions || [])
              .map((opt) => normalizeGroupValue(opt.value))
              .filter(Boolean);
            currentGroupFilter = new Set(selected);
            if (lastInspect) renderSubjectPreview(lastInspect);
          }});
        }}

        function setStatus(text, color='sage') {{
          statusBadge.textContent = text;
          statusBadge.classList.remove('hidden');
          statusBadge.className = `inline-flex items-center px-2 py-1 text-xs rounded-full bg-${{color}}-50 text-${{color}}-700 ring-1 ring-${{color}}-200`;
        }}

        function escapeHtml(str) {{
          return String(str)
            .replace(/&/g, '&amp;')
            .replace(/</g, '&lt;')
            .replace(/>/g, '&gt;')
            .replace(/"/g, '&quot;')
            .replace(/'/g, '&#39;');
        }}

        function escapeSelector(str) {{
          if (window.CSS && CSS.escape) {{
            return CSS.escape(str);
          }}
          return String(str).replace(/"/g, '\\"');
        }}

        function closeUnassignedModal() {{
          if (unassignedModal) {{
            unassignedModal.classList.add('hidden');
          }}
        }}

        async function openUnassignedModal(subjectName) {{
          if (!unassignedModal || !subjectName) return;
          const jobId = latestPivotJobId || currentJobId;
          const safeName = String(subjectName);
          const targetSubjectKey = safeName.trim();
          if (unassignedTitle) unassignedTitle.textContent = `${{safeName}} - 미배정 학생`;
          if (unassignedCount) unassignedCount.textContent = '';
          if (unassignedContent) unassignedContent.innerHTML = '<div class=\"text-stone-500\">데이터를 불러오는 중...</div>';
          unassignedModal.classList.remove('hidden');
          if (!jobId) {{
            if (unassignedContent) unassignedContent.innerHTML = '<div class=\"text-stone-500\">최근 실행 기록이 없습니다.</div>';
            return;
          }}
          try {{
            const res = await fetch(`/jobs/${{jobId}}/unassigned/${{encodeURIComponent(safeName)}}`);
            if (!res.ok) throw new Error('조회 실패');
            const js = await res.json();
            const list = Array.isArray(js.students) ? js.students : [];
            if (unassignedCount) {{
              unassignedCount.textContent = list.length ? `미배정 ${{list.length}}명` : '미배정 없음';
            }}
            if (!unassignedContent) return;
            if (!list.length) {{
              unassignedContent.innerHTML = '<div class=\"text-stone-500\">미배정 학생이 없습니다.</div>';
              return;
            }}
            const rows = list.map((st, idx) => {{
              const sid = escapeHtml(st.student_id ?? st.id ?? '');
              const nm = escapeHtml(st.name ?? '');
              const choices = Array.isArray(st.choices) ? st.choices : [];
              const choiceBadges = choices.map((ch) => {{
                const rawSubj = (ch?.subject ?? '').toString();
                const subj = escapeHtml(rawSubj);
                const status = (ch?.status ?? '').toString();
                const slot = (ch?.slot ?? '').toString().trim();
                const badgeText = status === 'assigned' ? (slot || '배정') : '미배정';
                const badgeClass = status === 'assigned' ? 'bg-sage-100 text-sage-700' : 'bg-rose-100 text-rose-700';
                const highlight = rawSubj.trim() === targetSubjectKey ? 'font-semibold text-stone-900' : 'text-stone-700';
                return `<span class=\"inline-flex items-center gap-1 px-2 py-1 rounded-full border border-stone-200 bg-white text-xs\">
                  <span class=\"${{highlight}}\">${{subj}}</span>
                  <span class=\"px-1 rounded ${{badgeClass}}\">${{escapeHtml(badgeText)}}</span>
                </span>`;
              }}).join(' ');
              const choicesCell = choiceBadges || '<span class=\"text-stone-500\">-</span>';
              return `<tr class=\"odd:bg-white even:bg-sage-50\">
                <td class=\"px-2 py-1 text-right text-stone-500\">${{idx + 1}}</td>
                <td class=\"px-2 py-1 font-mono text-sm text-stone-800\">${{sid}}</td>
                <td class=\"px-2 py-1 text-sm text-stone-800\">${{nm || '-'}}</td>
                <td class=\"px-2 py-1 text-sm text-stone-800\"><div class=\"flex flex-wrap gap-1\">${{choicesCell}}</div></td>
              </tr>`;
            }}).join('');
            unassignedContent.innerHTML = `
              <div class=\"overflow-x-auto\">
                <table class=\"min-w-full text-xs\">
                  <thead class=\"bg-sage-100 text-stone-700\">
                    <tr>
                      <th class=\"px-2 py-1 text-right w-10\">#</th>
                      <th class=\"px-2 py-1 text-left\">학번</th>
                      <th class=\"px-2 py-1 text-left\">이름</th>
                      <th class=\"px-2 py-1 text-left\">선택 과목 (배정)</th>
                    </tr>
                  </thead>
                  <tbody>${{rows}}</tbody>
                </table>
              </div>`;
          }} catch (err) {{
            if (unassignedContent) {{
              unassignedContent.innerHTML = '<div class=\"text-rose-600 text-sm\">불러오는 중 오류가 발생했습니다.</div>';
            }}
          }}
        }}

        async function downloadExportXlsx(jobId) {{
          if (!jobId) throw new Error('작업 ID가 없습니다.');
          // Use the currently rendered pivot row order (manual/total) as the export column order.
          const order = Array.from(pivotTable?.querySelectorAll('tbody tr[data-subject-key]') || [])
            .map((tr) => {{
              const key = tr.getAttribute('data-subject-key') || '';
              try {{ return decodeURIComponent(key); }} catch (e) {{ return key; }}
            }})
            .filter((s) => s && s.trim());

          const res = await fetch(`/jobs/${{jobId}}/export`, {{
            method: 'POST',
            headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify({{ subject_order: order }}),
          }});
          if (!res.ok) {{
            let msg = '내보내기 실패';
            try {{
              const js = await res.json();
              msg = js.error || msg;
            }} catch (e) {{}}
            throw new Error(msg);
          }}
          const blob = await res.blob();
          const url = URL.createObjectURL(blob);
          const a = document.createElement('a');
          a.href = url;
          a.download = '배정결과.xlsx';
          document.body.appendChild(a);
          a.click();
          a.remove();
          setTimeout(() => URL.revokeObjectURL(url), 1000);
        }}

        function recommendedSections(count) {{
          const capVal = Math.max(1, parseInt(capInput?.value || '28', 10) || 28);
          const demand = Number(count || 0);
          if (demand <= 0) return 0;
          return Math.max(1, Math.ceil(demand / capVal));
        }}

        function clearDownloads() {{
          downloads.innerHTML = '';
          downloads.classList.add('hidden');
        }}

        function slotPastelClassFromCol(colName) {{
          // colName example: "A-1", "B-3" (pivotData.columns entries after Total)
          try {{
            const slot = String(colName || '').split('-', 1)[0].trim().toUpperCase();
            switch (slot) {{
              case 'A': return 'bg-amber-50';
              case 'B': return 'bg-sky-50';
              case 'C': return 'bg-rose-50';
              case 'D': return 'bg-emerald-50';
              case 'E': return 'bg-violet-50';
              case 'F': return 'bg-lime-50';
              default: return 'bg-sage-50';
            }}
          }} catch (e) {{
            return 'bg-sage-50';
          }}
        }}

        function drawPivot() {{
          const cols = pivotData.columns;
          const rawRows = pivotData.rows.slice();
          let rows = [];
          if (rowOrderMode === 'manual') {{
            if (!manualSubjectOrder || !manualSubjectOrder.length) {{
              manualSubjectOrder = loadStoredSubjectOrder();
            }}
            const rowBySubject = new Map(rawRows.map((r) => [String(r[0]), r]));
            const subjectsNow = rawRows.map((r) => String(r[0]));
            const subjectsSet = new Set(subjectsNow);
            const order = [];
            const seen = new Set();
            (Array.isArray(manualSubjectOrder) ? manualSubjectOrder : []).forEach((s) => {{
              const subj = String(s);
              if (!subj || seen.has(subj) || !subjectsSet.has(subj)) return;
              seen.add(subj);
              order.push(subj);
            }});
            subjectsNow.forEach((subj) => {{
              if (!seen.has(subj)) {{
                seen.add(subj);
                order.push(subj);
              }}
            }});
            manualSubjectOrder = order.slice();
            saveStoredSubjectOrder(manualSubjectOrder);
            rows = order.map((subj) => rowBySubject.get(subj)).filter(Boolean);
          }} else {{
            if (pivotSortMode === 'total') {{
              rows = rawRows.sort((a, b) => {{
                const va = Number(a[1] || 0);
                const vb = Number(b[1] || 0);
                return sortAsc ? (va - vb) : (vb - va);
              }});
            }} else {{
              const normSubj = (s) => String(s ?? '').trim().replace(/\\s+/g, ' ');
              rows = rawRows.sort((a, b) => {{
                const cmp = normSubj(a[0]).localeCompare(normSubj(b[0]), 'ko', {{ numeric: true, sensitivity: 'base' }});
                return subjectSortAsc ? cmp : -cmp;
              }});
            }}
          }}
          // Build grouped headers
          const groups = pivotData.groups || [];
          const slotMeta = pivotData.slot_meta || {{}};
          const totalHeaderSuffix = rowOrderMode === 'manual' ? ' · 수동정렬' : '';
          const totalSortIndicator = (rowOrderMode === 'manual' || pivotSortMode !== 'total')
            ? ''
            : (' ' + (sortAsc ? '▲' : '▼'));
          const subjectSortIndicator = (rowOrderMode === 'manual' || pivotSortMode !== 'subject')
            ? ''
            : (' ' + (subjectSortAsc ? '▲' : '▼'));
          let hTop = '<tr>' +
            '<th id="th-subject" class="px-3 py-2 text-left text-stone-700 border-b border-stone-200 cursor-pointer select-none" rowspan="2">Subject' + subjectSortIndicator + '</th>' +
            '<th id="th-total" class="px-3 py-2 text-right text-stone-700 border-b border-stone-200 cursor-pointer select-none" rowspan="2">총인원(미배정)' + totalSortIndicator + totalHeaderSuffix + '</th>';
          for (const g of groups) {{
            const meta = slotMeta[g.slot] || {{}};
            const opened = Number(meta.sections_open || 0).toLocaleString();
            hTop += `<th class=\"px-3 py-2 text-center text-stone-700 border-b border-stone-200\" colspan=\"${{g.sections.length}}\">` +
                    `<div class=\"font-medium\">${{g.slot}}</div>` +
                    `<div class=\"text-xs text-stone-500\">총 ${{opened}}반</div>` +
                    `</th>`;
          }}
          hTop += '</tr>';
          let hSub = '<tr>';
          for (const g of groups) {{
            for (const n of g.sections) {{
              hSub += `<th class=\"px-3 py-1 text-right text-stone-600 border-b border-stone-200\">${{n}}</th>`;
            }}
          }}
          hSub += '</tr>';
          const trs = rows.map(r => {{
            const subject = String(r[0]);
            const subjectKey = encodeURIComponent(subject);
            const total = Number(r[1] || 0);
            const ua = (pivotData.row_meta && pivotData.row_meta[subject]) ? Number(pivotData.row_meta[subject].unassigned || 0) : 0;
            const safeSubject = escapeHtml(subject);
            let cells = '';
            // Subject with gear icon
            const hasConstraints = subjectConstraints[subject] && subjectConstraints[subject].maxPerSlot;
            const constraintIndicator = hasConstraints ?
              '<span class="inline-block w-2 h-2 bg-sage-500 rounded-full ml-1" title="제약 조건 설정됨"></span>' : '';
            cells += `<td class="px-3 py-1 text-left border-b border-stone-100">
              <div class="flex items-center gap-2">
                <span class="subject-drag-handle text-stone-400 hover:text-stone-600 select-none cursor-move" draggable="true" data-subject-key="${{subjectKey}}" title="드래그하여 과목 순서 변경">⋮⋮</span>
                <span>${{safeSubject}}</span>
                ${{constraintIndicator}}
                <button data-subject="${{safeSubject}}" class="constraint-btn text-stone-400 hover:text-stone-600" title="제약 조건 설정">
                  <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M10.325 4.317c.426-1.756 2.924-1.756 3.35 0a1.724 1.724 0 002.573 1.066c1.543-.94 3.31.826 2.37 2.37a1.724 1.724 0 001.065 2.572c1.756.426 1.756 2.924 0 3.35a1.724 1.724 0 00-1.066 2.573c.94 1.543-.826 3.31-2.37 2.37a1.724 1.724 0 00-2.572 1.065c-.426 1.756-2.924 1.756-3.35 0a1.724 1.724 0 00-2.573-1.066c-1.543.94-3.31-.826-2.37-2.37a1.724 1.724 0 00-1.065-2.572c-1.756-.426-1.756-2.924 0-3.35a1.724 1.724 0 001.066-2.573c-.94-1.543.826-3.31 2.37-2.37.996.608 2.296.07 2.572-1.065z"></path>
                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M15 12a3 3 0 11-6 0 3 3 0 016 0z"></path>
                  </svg>
                </button>
              </div>
            </td>`;
            // Total (Unassigned)
            const uaBadge = ua > 0
              ? `<button type="button" class="unassigned-btn text-rose-600 underline decoration-dotted hover:text-rose-700 focus:outline-none cursor-pointer" data-subject="${{safeSubject}}" title="미배정 학번 보기">${{ua.toLocaleString()}}</button>`
              : `<span class="text-stone-500">${{ua.toLocaleString()}}</span>`;
            cells += `<td class=\"px-3 py-1 text-right border-b border-stone-100\">${{total.toLocaleString()}} (${{uaBadge}})</td>`;
            // Section counts start at index 2
            for (let i = 2; i < r.length; i++) {{
              const v = Number(r[i] || 0);
              const colName = cols[i];
              const pastel = v > 0 ? slotPastelClassFromCol(colName) : '';
              cells += `<td class=\"px-3 py-1 text-right border-b border-stone-100 ${{pastel}}\">${{v.toLocaleString()}}</td>`;
            }}
            return `<tr class=\"odd:bg-white even:bg-sage-50\" data-subject-key=\"${{subjectKey}}\">${{cells}}</tr>`;
          }}).join('');
          pivotTable.innerHTML = `
            <thead class=\"bg-sage-100 sticky top-0\">${{hTop}}${{hSub}}</thead>
            <tbody>${{trs}}</tbody>
          `;
          pivotWrap.classList.remove('hidden');
          initPivotDnD();
          // Attach sorter
          const thSubject = document.getElementById('th-subject');
          if (thSubject) {{
            thSubject.onclick = () => {{
              rowOrderMode = 'total';
              pivotSortMode = 'subject';
              subjectSortAsc = !subjectSortAsc;
              drawPivot();
            }};
            thSubject.title = '과목명 정렬 토글 (클릭 시 수동정렬 해제)';
          }}
          const thTotal = document.getElementById('th-total');
          if (thTotal) {{
            thTotal.onclick = () => {{
              rowOrderMode = 'total';
              pivotSortMode = 'total';
              sortAsc = !sortAsc;
              drawPivot();
            }};
            thTotal.title = '총인원 정렬 (클릭 시 수동정렬 해제)';
          }}
        }}

        function renderPivot(pivot) {{
          pivotData = pivot;
          if (!pivot || !pivot.columns || !pivot.rows) {{
            pivotWrap.classList.add('hidden');
            pivotTable.innerHTML = '';
            return;
          }}
          // Load saved manual order for later drag/reorder, but keep default view sorting by subject.
          const stored = loadStoredSubjectOrder();
          if (stored && stored.length) manualSubjectOrder = stored.slice();
          drawPivot();
        }}

        // Preview with optional group filtering
        function renderSubjectPreview(info) {{
          try {{
            const subjects = Array.isArray(info?.subjects) ? info.subjects : [];
            const hasGroupFilter = currentGroupFilter && currentGroupFilter.size > 0;
            const filtered = hasGroupFilter
              ? subjects.filter((s) => currentGroupFilter.has(normalizeGroupValue(s?.group)))
              : subjects;
            const list = filtered.length ? filtered : subjects;
            if (!list.length) {{
              subjectPreviewBody.innerHTML = '<tr><td class=\"px-3 py-3 text-sm text-stone-500\">표시할 과목이 없습니다.</td></tr>';
              subjectPreviewWrap.classList.remove('hidden');
              return;
            }}
            subjectCounts = {{}};
            const rows = list.map((s) => {{
              const name = (s?.name ?? '').toString().trim();
              if (!name) return '';
              const count = Number(s?.count ?? 0);
              subjectCounts[name] = count;
              const recommended = recommendedSections(count);
              if (!(name in fixedSectionTargets)) {{
                fixedSectionTargets[name] = recommended > 0 ? recommended : null;
              }}
              const lockedValue = fixedSectionTargets[name];
              const displayVal = (lockedValue === 0 || lockedValue) ? lockedValue : '';
              const effectiveSections = displayVal !== '' ? Number(displayVal) : (recommended > 0 ? recommended : 0);
              const averageText = effectiveSections > 0 ? (count / effectiveSections).toFixed(1) : '-';
              const safeName = escapeHtml(name);
              const hasOtherConstraints = subjectConstraints[name] && subjectConstraints[name].maxPerSlot;
              const indicator = hasOtherConstraints ? '<span class="inline-block w-2 h-2 bg-sage-500 rounded-full" title="기타 제약 있음"></span>' : '';
              const groupLabel = (s?.group ?? '').toString().trim();
              const groupBadge = groupLabel ? `<span class="text-xs px-2 py-0.5 rounded-full bg-sage-100 text-sage-700">${{escapeHtml(groupLabel)}}</span>` : '';
              return `
                <tr class="odd:bg-white even:bg-sage-50">
                  <td class="px-3 py-2 align-top">
                    <div class="flex items-center gap-2 flex-wrap">
                      <span class="font-medium text-stone-800 truncate" title="${{safeName}}">${{safeName}}</span>
                      ${{groupBadge}}
                      ${{indicator}}
                      <button data-subject="${{safeName}}" class="constraint-btn text-stone-400 hover:text-stone-600" title="제약 조건 설정">
                        <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M10.325 4.317c.426-1.756 2.924-1.756 3.35 0a1.724 1.724 0 002.573 1.066c1.543-.94 3.31.826 2.37 2.37a1.724 1.724 0 001.065 2.572c1.756.426 1.756 2.924 0 3.35a1.724 1.724 0 00-1.066 2.573c.94 1.543-.826 3.31-2.37 2.37a1.724 1.724 0 00-2.572 1.065c-.426 1.756-2.924 1.756-3.35 0a1.724 1.724 0 00-2.573-1.066c-1.543.94-3.31-.826-2.37-2.37a1.724 1.724 0 00-1.065-2.572c-1.756-.426-1.756-2.924 0-3.35a1.724 1.724 0 001.066-2.573c-.94-1.543.826-3.31 2.37-2.37.996.608 2.296.07 2.572-1.065z"></path>
                          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M15 12a3 3 0 11-6 0 3 3 0 016 0z"></path>
                        </svg>
                      </button>
                    </div>
                    <div class="text-xs text-stone-500 mt-0.5">선택 ${{count.toLocaleString()}}명</div>
                  </td>
                  <td class="px-3 py-2 text-right align-top">${{count.toLocaleString()}}</td>
                  <td class="px-3 py-2 text-right align-top">${{recommended}}</td>
                  <td class="px-3 py-2 align-top">
                    <div class="flex items-center gap-2">
                      <input type="number" min="0" class="section-input w-20 rounded-md border border-stone-300 px-2 py-1 text-right"
                        data-subject="${{safeName}}" data-recommended="${{recommended}}" value="${{displayVal === '' ? '' : displayVal}}"
                        placeholder="${{recommended || ''}}" />
                      <button type="button" class="section-reset text-xs text-stone-500 hover:text-stone-700"
                        data-subject="${{safeName}}" data-recommended="${{recommended}}">↺</button>
                    </div>
                    <p class="mt-1 text-xs text-stone-400">비워두면 자동 결정</p>
                  </td>
                  <td class="px-3 py-2 text-right align-top">
                    <span class="avg-text" data-subject="${{safeName}}" data-count="${{count}}" data-recommended="${{recommended}}">${{averageText}}</span>
                  </td>
                </tr>`;
            }}).filter(Boolean).join('');
            subjectPreviewBody.innerHTML = rows;
            subjectPreviewWrap.classList.remove('hidden');
            subjectPreviewBody.querySelectorAll('.section-input').forEach((inp) => {{
              inp.addEventListener('input', () => handleSectionInput(inp));
            }});
            subjectPreviewBody.querySelectorAll('.section-reset').forEach((btn) => {{
              btn.addEventListener('click', () => resetSectionValue(btn));
            }});
          }} catch (err) {{
            console.warn('preview render error', err);
            subjectPreviewBody.innerHTML = '';
            subjectPreviewWrap.classList.add('hidden');
          }}
        }}

        if (applySectionRecommendBtn) {{
          applySectionRecommendBtn.addEventListener('click', () => {{
            if (!lastInspect) return;
            const subjects = Array.isArray(lastInspect.subjects) ? lastInspect.subjects : [];
            subjects.forEach((s) => {{
              const name = (s?.name ?? '').toString().trim();
              if (!name) return;
              const rec = recommendedSections(Number(s?.count ?? 0));
              fixedSectionTargets[name] = rec > 0 ? rec : null;
            }});
            renderSubjectPreview(lastInspect);
          }});
        }}
        if (releaseAllSectionsBtn) {{
          releaseAllSectionsBtn.addEventListener('click', () => {{
            Object.keys(fixedSectionTargets).forEach((key) => {{
              fixedSectionTargets[key] = null;
            }});
            if (lastInspect) renderSubjectPreview(lastInspect);
          }});
        }}

        function handleSectionInput(inputEl) {{
          const subj = inputEl?.dataset?.subject;
          if (!subj) return;
          const raw = inputEl.value.trim();
          if (raw === '') {{
            fixedSectionTargets[subj] = null;
            updateAverageDisplay(subj);
            return;
          }}
          let val = parseInt(raw, 10);
          if (!Number.isFinite(val)) {{
            fixedSectionTargets[subj] = null;
            updateAverageDisplay(subj);
            return;
          }}
          if (val < 0) val = 0;
          fixedSectionTargets[subj] = val;
          updateAverageDisplay(subj);
        }}

        function resetSectionValue(btn) {{
          const subj = btn?.dataset?.subject;
          if (!subj) return;
          const rec = parseInt(btn.dataset.recommended || '0', 10);
          fixedSectionTargets[subj] = rec > 0 ? rec : null;
          const target = Array.from(subjectPreviewBody.querySelectorAll('.section-input')).find((el) => el.dataset.subject === subj);
          if (target) {{
            target.value = rec > 0 ? rec : '';
          }}
          updateAverageDisplay(subj);
        }}

        function updateAverageDisplay(subjectName) {{
          const sel = escapeSelector(subjectName);
          const avgEl = subjectPreviewBody.querySelector(`.avg-text[data-subject=\"${{sel}}\"]`);
          if (!avgEl) return;
          const count = Number(avgEl.dataset.count || subjectCounts[subjectName] || 0);
          const recommended = Number(avgEl.dataset.recommended || 0);
          const input = subjectPreviewBody.querySelector(`.section-input[data-subject=\"${{sel}}\"]`);
          let sections = null;
          if (input) {{
            const val = input.value.trim();
            if (val !== '') sections = Number(val);
          }}
          if (sections === null || !Number.isFinite(sections) || sections <= 0) {{
            sections = recommended > 0 ? recommended : null;
          }}
          if (sections && sections > 0) {{
            avgEl.textContent = (count / sections).toFixed(1);
          }} else {{
            avgEl.textContent = '-';
          }}
        }}

        // Inspect uploaded file for quick preview (groups, semesters, subjects, headcount)
        async function inspectFile(file) {{
          const fd = new FormData();
          fd.append('xlsx', file);
          const res = await fetch('/inspect', {{ method: 'POST', body: fd }});
          if (!res.ok) return null;
          const js = await res.json();
          return js;
        }}

        inputFile.addEventListener('change', async (e) => {{
          const f = e.target.files && e.target.files[0];
          if (groupSel) {{
            groupSel.innerHTML = '';
            groupSel.multiple = false;
            groupSel.size = 1;
          }}
          groupRow.classList.add('hidden');
          applyGroupFilter([]);
          fixedSectionTargets = {{}};
          subjectPreviewBody.innerHTML = '';
          subjectPreviewWrap.classList.add('hidden');
          if (!f) return;
          try {{
            const info = await inspectFile(f);
            const groups = (info && Array.isArray(info.groups)) ? info.groups : [];
            // Suggest rooms by class_count
            const roomsInp = form.querySelector('input[name="rooms"]');
            if (roomsInp && Number(info?.class_count||0) > 0) roomsInp.value = String(Number(info.class_count));
            // Cache and render preview
            lastInspect = info;
            renderSubjectPreview(lastInspect);
            const enableLegacyPreview = false;
            // Inline preview (kept for backward-compat; ensure braces escaped in f-string)
            if (enableLegacyPreview) {{
            try {{
              const subs = Array.isArray(info?.subjects) ? info.subjects : [];
              const slots = Number(form.querySelector('input[name="slots"]').value || 4);
              const labels = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'.slice(0, Math.max(0, slots)).split('');
              if (subs.length) {{
                let thead = '<tr>' +
                  '<th class="px-3 py-2 text-left text-stone-700 border-b border-stone-200">Subject</th>' +
                  '<th class="px-3 py-2 text-right text-stone-700 border-b border-stone-200">총인원</th>';
                for (const lb of labels) {{ thead += `<th class=\"px-3 py-2 text-right text-stone-700 border-b border-stone-200\">${{lb}}</th>`; }}
                thead += '</tr>';
                const rows = subs.map(s => {{
                  const total = Number(s.count||0).toLocaleString();
              const hasConstraints = subjectConstraints[s.name] && subjectConstraints[s.name].maxPerSlot;
                  const constraintIndicator = hasConstraints ?
                    '<span class="inline-block w-2 h-2 bg-sage-500 rounded-full ml-1" title="제약 조건 설정됨"></span>' : '';
                  let cells = '';
                  cells += `<td class="px-3 py-1 text-left border-b border-stone-100">
                    <div class="flex items-center gap-2">
                      <span>${{s.name}}</span>
                      ${{constraintIndicator}}
                      <button data-subject="${{s.name}}" class="constraint-btn text-stone-400 hover:text-stone-600" title="제약 조건 설정">
                        <svg class="w-4 h-4" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M10.325 4.317c.426-1.756 2.924-1.756 3.35 0a1.724 1.724 0 002.573 1.066c1.543-.94 3.31.826 2.37 2.37a1.724 1.724 0 001.065 2.572c1.756.426 1.756 2.924 0 3.35a1.724 1.724 0 00-1.066 2.573c.94 1.543-.826 3.31-2.37 2.37a1.724 1.724 0 00-2.572 1.065c-.426 1.756-2.924 1.756-3.35 0a1.724 1.724 0 00-2.573-1.066c-1.543.94-3.31-.826-2.37-2.37a1.724 1.724 0 00-1.065-2.572c-1.756-.426-1.756-2.924 0-3.35a1.724 1.724 0 001.066-2.573c-.94-1.543.826-3.31 2.37-2.37.996.608 2.296.07 2.572-1.065z"></path>
                          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M15 12a3 3 0 11-6 0 3 3 0 016 0z"></path>
                        </svg>
                      </button>
                    </div>
                  </td>`;
                  cells += `<td class=\"px-3 py-1 text-right border-b border-stone-100\">${{total}}</td>`;
                  for (const _ of labels) {{ cells += `<td class=\"px-3 py-1 text-right border-b border-stone-100\">0</td>`; }}
                  return `<tr class=\"odd:bg-white even:bg-sage-50\">${{cells}}</tr>`;
                }}).join('');
                pivotTable.innerHTML = `<thead class=\"bg-sage-100\">${{thead}}</thead><tbody>${{rows}}</tbody>`;
                pivotWrap.classList.remove('hidden');
              }}
            }} catch (err) {{ console.warn('preview render error', err); }}
            }}
            // Render preview summary
            inspectBody.innerHTML = '';
            inspectWrap.classList.add('hidden');
            // Group selection
            if (groupSel) {{
              groupSel.innerHTML = '';
              if (groups.length > 1) {{
                groupSel.multiple = true;
                groupSel.size = Math.min(groups.length, 6);
                for (const g of groups) {{
                  const opt = document.createElement('option');
                  opt.value = String(g);
                  opt.textContent = String(g);
                  groupSel.appendChild(opt);
                }}
                applyGroupFilter([]);
                groupRow.classList.remove('hidden');
              }} else if (groups.length === 1) {{
                groupSel.multiple = false;
                groupSel.size = 1;
                const opt = document.createElement('option');
                opt.value = String(groups[0]);
                opt.textContent = String(groups[0]);
                opt.selected = true;
                groupSel.appendChild(opt);
                applyGroupFilter([groups[0]]);
                groupRow.classList.add('hidden');
              }} else {{
                groupSel.multiple = false;
                groupSel.size = 1;
                applyGroupFilter([]);
                groupRow.classList.add('hidden');
              }}
            }}
            renderSubjectPreview(lastInspect);
          }} catch (err) {{
            console.warn('inspect error', err);
            inspectBody.innerHTML = '';
            inspectWrap.classList.add('hidden');
            subjectPreviewBody.innerHTML = '';
            subjectPreviewWrap.classList.add('hidden');
          }}
        }});

        const slotsInp = form.querySelector('input[name="slots"]');
        const rerenderPreview = () => {{ if (lastInspect) renderSubjectPreview(lastInspect); }};
        if (slotsInp) {{
          slotsInp.addEventListener('input', rerenderPreview);
          slotsInp.addEventListener('change', rerenderPreview);
        }}
        if (capInput) {{
          capInput.addEventListener('input', rerenderPreview);
          capInput.addEventListener('change', rerenderPreview);
        }}

        // Constraint modal functions
        const constraintModal = document.getElementById('constraint-modal');
        const modalTitle = document.getElementById('modal-subject-title');
        const maxPerSlotInput = document.getElementById('max-per-slot');
        const closeModalBtn = document.getElementById('close-modal');
        const clearConstraintsBtn = document.getElementById('clear-constraints');
        const saveConstraintsBtn = document.getElementById('save-constraints');

        function openConstraintModal(subjectName) {{
          currentModalSubject = subjectName;
          modalTitle.textContent = `${{subjectName}} - 제약 조건 설정`;

          // Load existing constraints
          const existing = subjectConstraints[subjectName] || {{}};
          maxPerSlotInput.value = existing.maxPerSlot || '';

          constraintModal.classList.remove('hidden');
        }}

        function closeConstraintModal() {{
          constraintModal.classList.add('hidden');
          currentModalSubject = null;
        }}

        function saveConstraints() {{
          if (!currentModalSubject) return;

          const maxPerSlot = maxPerSlotInput.value.trim();

          if (!maxPerSlot) {{
            // No constraints - remove from object
            delete subjectConstraints[currentModalSubject];
          }} else {{
            const constraints = {{}};
            if (maxPerSlot) constraints.maxPerSlot = parseInt(maxPerSlot);
            subjectConstraints[currentModalSubject] = constraints;
          }}

          // Re-render current table to show constraint indicators
          if (lastInspect) {{
            renderSubjectPreview(lastInspect);
          }} else if (pivotData) {{
            drawPivot();
          }}

          closeConstraintModal();
        }}

        function clearConstraints() {{
          if (!currentModalSubject) return;
          delete subjectConstraints[currentModalSubject];

          // Re-render current table
          if (lastInspect) {{
            renderSubjectPreview(lastInspect);
          }} else if (pivotData) {{
            drawPivot();
          }}

          closeConstraintModal();
        }}

        // Modal event listeners
        closeModalBtn.addEventListener('click', closeConstraintModal);
        saveConstraintsBtn.addEventListener('click', saveConstraints);
        clearConstraintsBtn.addEventListener('click', clearConstraints);
        if (closeUnassignedBtn) closeUnassignedBtn.addEventListener('click', closeUnassignedModal);

        // Close modal when clicking outside
        constraintModal.addEventListener('click', (e) => {{
          if (e.target === constraintModal) closeConstraintModal();
        }});
        if (unassignedModal) {{
          unassignedModal.addEventListener('click', (e) => {{
            if (e.target === unassignedModal) closeUnassignedModal();
          }});
        }}

        // Close modal on Escape key
        document.addEventListener('keydown', (e) => {{
          if (e.key === 'Escape') {{
            if (constraintModal && !constraintModal.classList.contains('hidden')) {{
              closeConstraintModal();
            }}
            if (unassignedModal && !unassignedModal.classList.contains('hidden')) {{
              closeUnassignedModal();
            }}
          }}
        }});

        // Add event delegation for constraint buttons
        document.addEventListener('click', (e) => {{
          if (e.target.closest('.export-btn')) {{
            const btn = e.target.closest('.export-btn');
            const jobId = btn.getAttribute('data-job') || (latestPivotJobId || currentJobId);
            btn.disabled = true;
            btn.classList.add('opacity-60', 'cursor-not-allowed');
            const prev = statusNote.textContent;
            statusNote.textContent = '배정결과.xlsx 생성 중...';
            downloadExportXlsx(jobId)
              .catch((err) => {{
                console.warn('export error', err);
                errBox.textContent = String(err);
                errBox.classList.remove('hidden');
              }})
              .finally(() => {{
                btn.disabled = false;
                btn.classList.remove('opacity-60', 'cursor-not-allowed');
                statusNote.textContent = prev;
              }});
            return;
          }}
          if (e.target.closest('.unassigned-btn')) {{
            const btn = e.target.closest('.unassigned-btn');
            const subject = btn.getAttribute('data-subject');
            if (subject) {{
              openUnassignedModal(subject);
            }}
            return;
          }}
          if (e.target.closest('.constraint-btn')) {{
            const btn = e.target.closest('.constraint-btn');
            const subject = btn.getAttribute('data-subject');
            if (subject) {{
              openConstraintModal(subject);
            }}
          }}
        }});

        // Make functions globally available (for backward compatibility)
        window.openConstraintModal = openConstraintModal;

        form.addEventListener('submit', async (e) => {{
          e.preventDefault();
          clearDownloads();
          errBox.classList.add('hidden');
          statusNote.textContent = '업로드 중...';
          setStatus('PENDING');
          submitBtn.disabled = true;
          try {{
            const fd = new FormData(form);

            // Add constraint data as a hidden input field instead of direct FormData append
            console.log('[DEBUG] Current subjectConstraints:', subjectConstraints);
            const constraintsJsonStr = Object.keys(subjectConstraints).length > 0 ?
              JSON.stringify(subjectConstraints) : '{{}}';
            console.log('[DEBUG] Sending constraints JSON:', constraintsJsonStr);
            console.log('[DEBUG] About to submit form...');

            // Create a hidden input field for constraints
            const constraintsInput = document.createElement('input');
            constraintsInput.type = 'hidden';
            constraintsInput.name = 'constraints_json';
            constraintsInput.value = constraintsJsonStr;
            form.appendChild(constraintsInput);

            const sectionTotalsInput = document.createElement('input');
            sectionTotalsInput.type = 'hidden';
            sectionTotalsInput.name = 'section_totals_json';
            const sectionPayload = {{}};
            Object.entries(fixedSectionTargets).forEach(([subject, value]) => {{
              if (value === null || value === undefined || value === '') return;
              const num = parseInt(value, 10);
              if (!Number.isFinite(num)) return;
              sectionPayload[subject] = num;
            }});
            sectionTotalsInput.value = JSON.stringify(sectionPayload);
            form.appendChild(sectionTotalsInput);

            // Recreate FormData to include the new input
            const newFd = new FormData(form);

            // Remove the temporary input
            form.removeChild(constraintsInput);
            form.removeChild(sectionTotalsInput);

            console.log('[DEBUG] Sending POST request to /run...');
            const res = await fetch('/run', {{ method: 'POST', body: newFd }});
            console.log('[DEBUG] Received response:', res.status, res.statusText);
            const data = await res.json();
            if (!res.ok) throw new Error(data.error || '실행 실패');
            currentJobId = data.job;
            jobInfo.classList.remove('hidden');
            jobInfo.textContent = `작업 ID: ${{data.job}}`;
            statusNote.textContent = '대기열에 추가되었습니다. 처리 중...';
            if (pollTimer) clearInterval(pollTimer);
            pollTimer = setInterval(async () => {{
              const st = await fetch(`/jobs/${{data.job}}`);
              const js = await st.json();
              if (js.status === 'RUNNING') {{
                setStatus('RUNNING');
                if (mascot) mascot.classList.remove('hidden');
                if (mascotVideo) mascotVideo.classList.remove('hidden');
                if (mascotDone) mascotDone.classList.add('hidden');
                if (js.progress) {{ lastProgress = js.progress; }}
                if (runStartMs === null) {{
                  runStartMs = Date.now();
                  if (tickTimer) {{ clearInterval(tickTimer); tickTimer = null; }}
                  tickTimer = setInterval(() => {{
                    const secs = ((Date.now() - runStartMs) / 1000).toFixed(1);
                    const sols = (lastProgress && lastProgress.solutions !== undefined) ? Number(lastProgress.solutions||0).toLocaleString() : '0';
                    const ua = (lastProgress && lastProgress.unassigned !== undefined) ? ', 미배정 ' + Number(lastProgress.unassigned||0).toLocaleString() : '';
                    statusNote.textContent = `계산 중 · ${{secs}}s · 솔루션 ${{sols}}${{ua}}`;
                  }}, 200);
                }} else if (js.progress) {{
                  // refresh cached numbers even if timer already running
                  lastProgress = js.progress;
                }}
                // Immediate text update (before next tick)
                const secsNow = ((runStartMs ? (Date.now() - runStartMs) : 0) / 1000).toFixed(1);
                const solsNow = (lastProgress && lastProgress.solutions !== undefined) ? Number(lastProgress.solutions||0).toLocaleString() : '0';
                const uaNow = (lastProgress && lastProgress.unassigned !== undefined) ? ', 미배정 ' + Number(lastProgress.unassigned||0).toLocaleString() : '';
                statusNote.textContent = `계산 중 · ${{secsNow}}s · 솔루션 ${{solsNow}}${{uaNow}}`;
              }} else if (js.status === 'DONE') {{
                setStatus('DONE');
                if (mascot) mascot.classList.remove('hidden');
                if (mascotVideo) mascotVideo.classList.add('hidden');
                if (mascotDone) mascotDone.classList.remove('hidden');
                if (tickTimer) {{ clearInterval(tickTimer); tickTimer = null; }}
                runStartMs = null; lastProgress = null;
                if (js.summary && js.summary.total_unassigned !== undefined) {{
                  const ua = Number(js.summary.total_unassigned||0).toLocaleString();
                  statusNote.textContent = `미배정 ${{ua}}명으로 최종 완료했습니다.`;
                }} else {{
                  statusNote.textContent = '완료되었습니다. 아래 파일을 다운로드하세요.';
                }}
                clearInterval(pollTimer);
                downloads.classList.remove('hidden');
                downloads.innerHTML = `
                  <a class=\"inline-flex items-center px-3 py-2 rounded-lg bg-sage-600 text-white hover:bg-sage-700\" href=\"${{js.sections}}\" download>sections_plan.csv</a>
                  <a class=\"inline-flex items-center px-3 py-2 rounded-lg bg-sage-600 text-white hover:bg-sage-700\" href=\"${{js.assignments}}\" download>assignments.csv</a>
                  <a class=\"inline-flex items-center px-3 py-2 rounded-lg bg-sage-600 text-white hover:bg-sage-700\" href=\"${{js.report}}\" download>report.txt</a>
                  <button type=\"button\" class=\"export-btn inline-flex items-center px-3 py-2 rounded-lg bg-sage-600 text-white hover:bg-sage-700\" data-job=\"${{data.job}}\">배정결과.xlsx ⭐</button>
                  ${{js.constraints ? `<a class=\\\"inline-flex items-center px-3 py-2 rounded-lg bg-sage-600 text-white hover:bg-sage-700\\\" href=\\\"${{js.constraints}}\\\" download>constraints.csv</a>` : ''}}
                `;
                latestPivotJobId = data.job;
                if (js.pivot) {{ renderPivot(js.pivot); }} else {{ renderPivot(null); }}
                submitBtn.disabled = false;
               }} else if (js.status === 'ERROR') {{
                setStatus('ERROR', 'rose');
                if (mascot) mascot.classList.add('hidden');
                if (mascotVideo) mascotVideo.classList.add('hidden');
                if (mascotDone) mascotDone.classList.add('hidden');
                if (tickTimer) {{ clearInterval(tickTimer); tickTimer = null; }}
                runStartMs = null; lastProgress = null;
                statusNote.textContent = '오류가 발생했습니다.';
                errBox.textContent = js.error || 'Unknown error';
                errBox.classList.remove('hidden');
                renderPivot(null);
                clearInterval(pollTimer);
                submitBtn.disabled = false;
               }} else if (js.status === 'PENDING') {{
                setStatus('PENDING');
                if (mascot) mascot.classList.add('hidden');
                if (mascotVideo) mascotVideo.classList.add('hidden');
                if (mascotDone) mascotDone.classList.add('hidden');
                if (tickTimer) {{ clearInterval(tickTimer); tickTimer = null; }}
                runStartMs = null; lastProgress = null;
                statusNote.textContent = '대기 중...';
              }}
            }}, 2000);
          }} catch (err) {{
            setStatus('ERROR', 'rose');
            statusNote.textContent = '요청 중 오류가 발생했습니다.';
            errBox.textContent = String(err);
            errBox.classList.remove('hidden');
            submitBtn.disabled = false;
          }}
        }});
      </script>
    </body>
    </html>
    """

# 아주 간단한 인메모리 상태 저장소 (서버 재시작 시 초기화되는 점만 유의)
JOBS = {}  # job_id -> {"status": "PENDING|RUNNING|DONE|ERROR", "dir": Path, "error": str|None}

def delete_file_safe(file_path: Path):
    """파일 안전 삭제"""
    try:
        if file_path.exists():
            file_path.unlink()
            print(f"[DELETE] Removed file: {file_path}")
    except Exception as e:
        print(f"[DELETE] Error removing file {file_path}: {e}")

def delete_job_directory(job_dir: Path):
    """작업 디렉토리 전체 삭제 (개인정보 보호)"""
    try:
        if job_dir.exists():
            shutil.rmtree(job_dir)
            print(f"[DELETE] Removed job directory: {job_dir}")
    except Exception as e:
        print(f"[DELETE] Error removing directory {job_dir}: {e}")

async def cleanup_job_folder(job_id: str, delay_seconds: int = 3600):
    """일정 시간 후 작업 폴더 자동 삭제"""
    await asyncio.sleep(delay_seconds)
    try:
        info = JOBS.get(job_id)
        if info and info.get("dir"):
            job_dir = info["dir"]
            if job_dir.exists():
                shutil.rmtree(job_dir)
                print(f"[AUTO-CLEANUP] Deleted job directory after {delay_seconds}s: {job_dir}")
    except Exception as e:
        print(f"[AUTO-CLEANUP] Error deleting job {job_id}: {e}")

async def run_optimizer(job_id: str, xlsx_path: Path, out_dir: Path,
                        slots: int, rooms: int, extra: int, cap: int, maxcap: int, group: str | None = None,
                        extra_total: int | None = None, constraints_csv_path: Path | None = None,
                        fixed_sections_csv_path: Path | None = None):
    JOBS[job_id]["status"] = "RUNNING"
    cmd = [
        "python", "optimize_student_sections.py",
        "--input", str(xlsx_path),
        "--output-dir", str(out_dir),
        "--slots", str(slots),
        "--rooms-per-slot", str(rooms),
        "--extra-rooms-per-slot", str(extra),
        "--cap", str(cap),
        "--maxcap", str(maxcap),
        "--time-limit", "240",
        "--workers", "8"  # 머신 코어/워커 수에 맞춰 조정
    ]
    if group:
        cmd += ["--group", str(group)]
    if extra_total is not None:
        cmd += ["--extra-total", str(extra_total)]
    if constraints_csv_path and constraints_csv_path.exists():
        cmd += ["--constraints-csv", str(constraints_csv_path)]
    if fixed_sections_csv_path and fixed_sections_csv_path.exists():
        cmd += ["--fixed-sections-csv", str(fixed_sections_csv_path)]
    # Debug: show launch command and whether constraints CSV attached
    try:
        print("[DEBUG] Launching optimizer:", " ".join(cmd))
        if constraints_csv_path:
            print(f"[DEBUG] constraints_csv_path: {constraints_csv_path} exists={constraints_csv_path.exists()}")
        if fixed_sections_csv_path:
            print(f"[DEBUG] fixed_sections_csv_path: {fixed_sections_csv_path} exists={fixed_sections_csv_path.exists()}")
    except Exception:
        pass
    # 비동기 실행
    proc = await asyncio.create_subprocess_exec(
        *cmd, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE
    )
    # Stream stdout to capture progress
    async def read_stdout():
        try:
            while True:
                line = await proc.stdout.readline()
                if not line:
                    break
                txt = line.decode(errors="ignore").strip()
                if txt.startswith("[PROGRESS]"):
                    import json as _json
                    try:
                        js = _json.loads(txt[len("[PROGRESS]"):].strip())
                        JOBS[job_id]["progress"] = js
                    except Exception:
                        pass
        except Exception:
            pass

    async def read_stderr():
        try:
            JOBS[job_id]["_stderr"] = b""
            while True:
                line = await proc.stderr.readline()
                if not line:
                    break
                JOBS[job_id]["_stderr"] += line
        except Exception:
            pass

    await asyncio.gather(read_stdout(), read_stderr())
    await proc.wait()
    if proc.returncode == 0:
        JOBS[job_id]["status"] = "DONE"
        # Build pivot for quick UI preview
        assignments_csv = out_dir / "assignments.csv"
        JOBS[job_id]["pivot"] = build_pivot(assignments_csv)
        # Compute simple summary metrics
        try:
            if assignments_csv.exists():
                df_asg = pd.read_csv(assignments_csv)
                total_unassigned = int((df_asg.get("status") == "unassigned").sum())
                students_with_unassigned = int(df_asg[df_asg.get("status") == "unassigned"]["student_id"].nunique())
                total_assigned = int((df_asg.get("status") == "assigned").sum())
                JOBS[job_id]["summary"] = {
                    "total_unassigned": total_unassigned,
                    "students_with_unassigned": students_with_unassigned,
                    "total_assigned": total_assigned,
                }
                # Cache unassigned student list by subject for quick UI lookup (even after file deletion)
                try:
                    ua_map = build_unassigned_lookup(df_asg)
                    if ua_map:
                        JOBS[job_id]["unassigned"] = ua_map
                except Exception:
                    pass
        except Exception:
            pass

        # 개인정보 보호: report.txt를 별도 보관 + 1시간 후 작업 폴더 자동 삭제
        try:
            report_src = out_dir / "report.txt"
            if report_src.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                report_dest = REPORTS_DIR / f"{timestamp}_{job_id}.txt"
                shutil.copy2(report_src, report_dest)
                print(f"[BACKUP] Report saved to: {report_dest}")
                # 1시간 후 작업 폴더 자동 삭제 (사용자가 다운로드할 시간 확보)
                asyncio.create_task(cleanup_job_folder(job_id, delay_seconds=3600))
        except Exception as e:
            print(f"[BACKUP] Error backing up report: {e}")
    else:
        JOBS[job_id]["status"] = "ERROR"
        err = JOBS[job_id].get("_stderr", b"")
        JOBS[job_id]["error"] = (err or b"").decode(errors="ignore")

@app.post("/run")
async def run(request: Request):
    # Parse form data manually
    form_data = await request.form()
    print(f"[DEBUG] All form data keys: {list(form_data.keys())}")

    xlsx = form_data.get("xlsx")
    slots = int(form_data.get("slots", 4))
    rooms = int(form_data.get("rooms", 7))
    extra = int(form_data.get("extra", 1))
    cap = int(form_data.get("cap", 28))
    maxcap = int(form_data.get("maxcap", 30))
    raw_group = form_data.get("group")
    extra_total = form_data.get("extra_total")
    constraints_json = form_data.get("constraints_json")
    section_totals_json = form_data.get("section_totals_json")

    print(f"[DEBUG] Received constraints_json: {constraints_json}")
    job_id = str(uuid.uuid4())
    job_dir = BASE / job_id; job_dir.mkdir(parents=True, exist_ok=True)
    out_dir = job_dir / "out"; out_dir.mkdir(exist_ok=True)
    xlsx_path = job_dir / "input.xlsx"
    with open(xlsx_path, "wb") as f:
        f.write(await xlsx.read())

    group_values = []
    try:
        group_values = [str(g).strip() for g in form_data.getlist("group") if str(g).strip()]
    except AttributeError:
        group_values = []
    if not group_values and raw_group:
        gval = str(raw_group).strip()
        if gval:
            group_values = [gval]
    group = ",".join(group_values) if group_values else None

    # Handle optional constraints from JSON
    constraints_csv_path = None
    if constraints_json:
        try:
            import json
            constraints_data = json.loads(constraints_json)
            print(f"[DEBUG] Received constraints JSON: {constraints_json}")
            print(f"[DEBUG] Parsed constraints data: {constraints_data}")

            if constraints_data:
                # Save constraints.csv under out_dir so it's downloadable via /download route
                constraints_csv_path = out_dir / "constraints.csv"

                # Convert JSON to CSV format (use csv.writer for proper newlines/quoting)
                with open(constraints_csv_path, "w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(["subject", "max_sections_per_slot", "max_total_sections"])
                    for subject, constraint in constraints_data.items():
                        max_per_slot = constraint.get("maxPerSlot")
                        max_total = constraint.get("maxTotal")
                        # Normalize None/empty to ''
                        mps = "" if max_per_slot is None or str(max_per_slot).strip() == "" else int(max_per_slot)
                        mt = "" if max_total is None or str(max_total).strip() == "" else int(max_total)
                        writer.writerow([str(subject), mps, mt])
                        print(f"[DEBUG] Added constraint: {subject} -> maxPerSlot={mps}, maxTotal={mt}")

                print(f"[DEBUG] Created constraints CSV at: {constraints_csv_path}")
        except Exception as e:
            print(f"[DEBUG] Error parsing constraints: {e}")
            pass  # Ignore constraint parsing errors

    fixed_sections_csv_path = None
    if section_totals_json:
        try:
            import json
            section_data = json.loads(section_totals_json)
            if section_data:
                fixed_sections_csv_path = out_dir / "fixed_sections.csv"
                with open(fixed_sections_csv_path, "w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(["subject", "total_sections"])
                    for subject, total in section_data.items():
                        try:
                            val = int(total)
                        except Exception:
                            continue
                        writer.writerow([str(subject), val])
                print(f"[DEBUG] Created fixed sections CSV at: {fixed_sections_csv_path}")
        except Exception as e:
            print(f"[DEBUG] Error parsing section totals: {e}")
            fixed_sections_csv_path = None

    JOBS[job_id] = {"status": "PENDING", "dir": job_dir, "error": None}
    # 백그라운드 태스크 시작 (즉시 응답)
    et_val = None
    try:
        if extra_total is not None and str(extra_total).strip() != "":
            et_val = int(str(extra_total).strip())
    except Exception:
        et_val = None
    asyncio.create_task(run_optimizer(job_id, xlsx_path, out_dir, slots, rooms, extra, cap, maxcap, group, et_val, constraints_csv_path, fixed_sections_csv_path))

    return {"job": job_id, "status_url": f"/jobs/{job_id}"}

@app.get("/jobs/{job_id}")
def job_status(job_id: str):
    info = JOBS.get(job_id)
    if not info:
        return JSONResponse(status_code=404, content={"error": "job not found"})
    resp = {"job": job_id, "status": info["status"]}
    if info.get("progress"):
        resp["progress"] = info["progress"]
    if info.get("progress"):
        resp["progress"] = info["progress"]
    if info.get("summary"):
        resp["summary"] = info["summary"]
        if info["status"] == "DONE":
            resp.update({
                "sections": f"/download/{job_id}/sections_plan.csv",
                "assignments": f"/download/{job_id}/assignments.csv",
                "report": f"/download/{job_id}/report.txt",
                "filled_template": f"/download/{job_id}/filled_template.xlsx",
            })
            # Expose constraints CSV if present (saved under out/)
            try:
                cspath = info["dir"] / "out" / "constraints.csv"
                if cspath.exists():
                    resp["constraints"] = f"/download/{job_id}/constraints.csv"
                fspath = info["dir"] / "out" / "fixed_sections.csv"
                if fspath.exists():
                    resp["section_totals"] = f"/download/{job_id}/fixed_sections.csv"
            except Exception:
                pass
            if info.get("pivot"):
                resp["pivot"] = info["pivot"]
    if info["status"] == "ERROR":
        resp["error"] = info["error"]
    return resp

@app.get("/jobs/{job_id}/unassigned/{subject}")
def job_unassigned(job_id: str, subject: str):
    """Return unassigned students (id + optional name) for a given subject in a finished job."""
    info = JOBS.get(job_id)
    if not info:
        return JSONResponse(status_code=404, content={"error": "job not found"})
    if info.get("status") != "DONE":
        return JSONResponse(status_code=400, content={"error": "job not completed"})

    ua_map = info.get("unassigned") or {}
    if ua_map:
        students = ua_map.get(subject, [])
        return {"subject": subject, "count": len(students), "students": students}

    # Fallback: if the file still exists, compute on demand
    try:
        job_dir = info.get("dir")
        if job_dir:
            assignments_path = Path(job_dir) / "out" / "assignments.csv"
            if assignments_path.exists():
                df = pd.read_csv(assignments_path)
                ua_map = build_unassigned_lookup(df)
                students = ua_map.get(subject, [])
                return {"subject": subject, "count": len(students), "students": students}
    except Exception:
        pass

    return {"subject": subject, "count": 0, "students": []}


@app.post("/jobs/{job_id}/export")
def job_export(job_id: str, background_tasks: BackgroundTasks, payload: dict = Body(default_factory=dict)):
    """Create a fresh export xlsx (new file, not modifying the input template)."""
    info = JOBS.get(job_id)
    if not info:
        return JSONResponse(status_code=404, content={"error": "job not found"})
    if info.get("status") != "DONE":
        return JSONResponse(status_code=400, content={"error": "job not completed"})

    subject_order = None
    try:
        raw = payload.get("subject_order")
        if isinstance(raw, list):
            subject_order = [str(x) for x in raw if str(x).strip()]
    except Exception:
        subject_order = None

    out_dir = info["dir"] / "out"
    assignments_path = out_dir / "assignments.csv"
    if not assignments_path.exists():
        return JSONResponse(status_code=404, content={"error": "assignments.csv not found (이미 다운로드되어 삭제되었을 수 있습니다)."})

    export_path = out_dir / "export.xlsx"
    try:
        build_export_xlsx(assignments_path, export_path, subject_order=subject_order)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"export build failed: {e}"})

    background_tasks.add_task(delete_file_safe, export_path)

    return FileResponse(
        export_path,
        filename="배정결과.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.get("/download/{job_id}/{name}")
def download(job_id: str, name: str, background_tasks: BackgroundTasks):
    info = JOBS.get(job_id)
    if not info: return JSONResponse(status_code=404, content={"error":"job not found"})
    path = info["dir"] / "out" / name

    # report.txt는 백업에서도 제공 가능
    if name == "report.txt" and not path.exists():
        # 작업 폴더가 삭제된 경우 백업된 report 찾기
        import glob
        pattern = str(REPORTS_DIR / f"*_{job_id}.txt")
        backup_files = glob.glob(pattern)
        if backup_files:
            path = Path(backup_files[0])
            print(f"[INFO] Serving report from backup: {path}")
        else:
            return JSONResponse(status_code=404, content={"error":"report file not found"})
    elif not path.exists():
        return JSONResponse(status_code=404, content={"error":"file not found"})

    # 개인정보 보호: 다운로드 후 해당 파일만 즉시 삭제
    # 전체 폴더는 1시간 후 자동 삭제됨
    if name == "assignments.csv":
        # 학생 개인정보 포함 파일 - 다운로드 후 즉시 삭제
        background_tasks.add_task(delete_file_safe, path)
        background_tasks.add_task(delete_file_safe, info["dir"] / "input.xlsx")
    elif name == "sections_plan.csv":
        # 과목별 통계 파일 - 다운로드 후 삭제
        background_tasks.add_task(delete_file_safe, path)
    elif name == "filled_template.xlsx":
        # 배정결과 파일 - 다운로드 후 삭제
        background_tasks.add_task(delete_file_safe, path)
    elif name == "report.txt":
        # report는 이미 백업됨 - 파일만 삭제 (백업된 파일은 유지)
        if path.parent.name == "out":  # 작업 폴더의 report만 삭제
            background_tasks.add_task(delete_file_safe, path)

    return FileResponse(path)

@app.post("/inspect")
async def inspect(xlsx: UploadFile):
    """Parse the uploaded Excel and return a preview with subject counts.
    Subject names are normalized and disambiguated exactly like the solver,
    so constraints set in the UI match the model's subject keys.
    """
    try:
        content = await xlsx.read()

        def norm(s):
            return " ".join(str(s).split())

        # Try original wide format first
        try:
            df = pd.read_excel(io.BytesIO(content), sheet_name=0)
            df.columns = [str(c).strip() for c in df.columns]
            looks_original = len(df.columns) >= 3 and (
                any("학번" in c or c.lower() in {"id", "student_id"} for c in df.columns[:2]) or
                any("이름" in c or c.lower() in {"name"} for c in df.columns[:2])
            )
        except Exception:
            df = pd.DataFrame(); looks_original = False

        if looks_original and not df.empty:
            id_col = df.columns[0]
            # Normalize subject column headers like the solver
            subject_cols = [norm(c) for c in df.columns[2:]]
            # Rename the dataframe columns so counts use normalized names
            df = df.rename(columns={old: new for old, new in zip(df.columns[2:], subject_cols)})
            # headcount = non-empty id rows
            headcount = int(df[id_col].notna().sum())
            # class_count from 5-digit student id: class = 3rd digit (index 2)
            classes = set()
            for v in df[id_col].dropna().astype(str):
                s = "".join(ch for ch in str(v) if ch.isdigit())
                if len(s) >= 3:
                    classes.add(s[2])
            subjects = []
            for c in subject_cols:
                col = df[c]
                try:
                    cnt = int((pd.to_numeric(col, errors='coerce').fillna(0.0) == 1).sum())
                except Exception:
                    cnt = int((col.astype(str).str.strip() == '1').sum())
            subjects.append({"name": c, "count": cnt, "group": ""})
            # Keep original Excel column order (no sorting)
            return {"groups": [], "semesters": [], "subjects": subjects, "headcount": headcount, "class_count": len(classes)}

        # New headered layout (with group/semester headers)
        df0 = pd.read_excel(io.BytesIO(content), sheet_name=0, header=None)
        if df0.shape[0] < 6 or df0.shape[1] < 5:
            return {"groups": [], "semesters": [], "subjects": [], "headcount": 0, "class_count": 0}
        COL_A, COL_B, COL_C, COL_E = 0, 1, 2, 4
        ROW_SEM, ROW_GRP, ROW_SUBJ, ROW_DATA = 1, 2, 3, 5
        sem_row = df0.iloc[ROW_SEM, COL_E:]
        grp_row = df0.iloc[ROW_GRP, COL_E:]
        subj_row = df0.iloc[ROW_SUBJ, COL_E:]

        subjects_names = []  # normalized + uniquely tagged names (or None)
        sems = []
        grps = []
        used = set()
        for j, base in enumerate(subj_row):
            if pd.isna(base):
                subjects_names.append(None)
                sems.append(None)
                grps.append(None)
                continue
            base_s = norm(base)
            if not base_s:
                subjects_names.append(None)
                sems.append(None)
                grps.append(None)
                continue
            sem = sem_row.iloc[j] if j < len(sem_row) else None
            grp = grp_row.iloc[j] if j < len(grp_row) else None
            # forward-fill merged header cells
            prev_sem = sems[-1] if sems else None
            prev_grp = grps[-1] if grps else None
            if (sem is None or (isinstance(sem, float) and pd.isna(sem)) or str(sem).strip() == ""):
                sem = prev_sem
            if (grp is None or (isinstance(grp, float) and pd.isna(grp)) or str(grp).strip() == ""):
                grp = prev_grp
            sem_s = str(sem).strip() if sem is not None and not pd.isna(sem) else ""
            grp_s = str(grp).strip() if grp is not None and not pd.isna(grp) else ""
            name = base_s
            # Ensure uniqueness if duplicated subject names appear across groups/semesters
            if name in used and (sem_s or grp_s):
                tag = "-".join([p for p in [sem_s, grp_s] if p])
                name = f"{base_s} [{tag}]"
            k = 2
            while name in used:
                name = f"{base_s} ({k})"
                k += 1
            used.add(name)
            subjects_names.append(name)
            sems.append(sem_s)
            grps.append(grp_s)

        # counts aligned with columns E..end
        subj_counts = []
        for j, sname in enumerate(subjects_names):
            if not sname:
                continue
            col = df0.iloc[ROW_DATA:, COL_E + j]
            try:
                cnt = int((pd.to_numeric(col, errors='coerce').fillna(0.0) == 1).sum())
            except Exception:
                cnt = int((col.astype(str).str.strip() == '1').sum())
            group_val = grps[j] if j < len(grps) and grps[j] else ""
            subj_counts.append({"name": sname, "count": cnt, "group": group_val})
        # Keep original Excel column order (no sorting)

        # headcount and class_count by student id across A/B/C
        headcount = 0
        classes = set()
        for i in range(ROW_DATA, df0.shape[0]):
            vals = []
            for ci in (COL_C, COL_B, COL_A):
                v = df0.iat[i, ci] if ci < df0.shape[1] else None
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    continue
                s = str(v).strip()
                if s == '' or s == '0':
                    continue
                vals.append(s)
            if vals:
                headcount += 1
                sid = vals[0]
                sd = "".join(ch for ch in str(sid) if ch.isdigit())
                if len(sd) >= 3:
                    classes.add(sd[2])
        uniq_groups = sorted({g for g in grps if g})
        uniq_sems = sorted({s for s in sems if s})
        return {"groups": uniq_groups, "semesters": uniq_sems, "subjects": subj_counts, "headcount": headcount, "class_count": len(classes)}
    except Exception:
        return {"groups": [], "semesters": [], "subjects": [], "headcount": 0, "class_count": 0}

# 실행: uvicorn app:app --host 0.0.0.0 --port 8000
